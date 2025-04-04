from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border

bold_font = Font(name='Arimo', size=10, bold=True)
regular_font = Font(name='Arimo', size=10)
thins = Side(border_style="thin")

wb = Workbook()
ws = wb.active

ws.column_dimensions['B'].width = 20
ws.column_dimensions['E'].width = 20
ws.column_dimensions['D'].width = 5

ws.row_dimensions[0].height = 70

ws['G1'] = "Т а б е л ь  №  ______________"
ws['G1'].font = bold_font
ws['F2'] = "учета использования рабочего времени"
ws['F2'].font = bold_font
ws['B5'] = "Учреждение"
ws['E5'] = "Ярославский государственный университет им. П.Г. Демидова       "
for x in range(5, 16):
    for y in range(5, 8):
        ws.cell(row=y, column=x).border = Border(bottom=thins)
ws['B6'] = "Структурное подразделение"
ws['E6'] = "Кафедра теоретической информатики                               "
ws['B7'] = "Вид табеля"
ws['E7'] = "первичный                                                       "
wb.save("sample.xlsx")
