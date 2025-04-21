import sys
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment

if len(sys.argv) != 2:
    exit("One argument is required!\n<date> in this format: 01/2020")

bold_font = Font(name='Arimo', size=10, bold=True)
regular_font = Font(name='Arimo', size=10)
small_font = Font(name='Arimo', size=8)
thins = Side(border_style="thin")

months = {
    "01" : "января",
    "02" : "февраля",
    "03" : "марта",
    "04" : "апреля",
    "05" : "мая",
    "06" : "июня",
    "07" : "июля",
    "08" : "августа",
    "09" : "сентября",
    "10" : "октября",
    "11" : "ноября",
    "12" : "декабря"}

wb = Workbook()
ws = wb.active
date = sys.argv[1].split("/")
days_amount = monthrange(int(date[1]), int(date[0]))[1]

def basic_formating():
    ws.row_dimensions[1].height = 70
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 18

    for i in range(1, 16):
        ws.column_dimensions[get_column_letter(5 + i)].width = 4
        ws[get_column_letter(5 + i) + "13"] = str(i)
    for i in range(1, days_amount - 14):
        ws.column_dimensions[get_column_letter(21 + i)].width = 4
        ws[get_column_letter(21 + i) + "13"] = str(15 + i)
    for i in range(1, days_amount + 2):
        ws[get_column_letter(5 + i) + "13"].alignment = Alignment(horizontal = "center")
    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(7 + days_amount + i)].width = 10
    for i in range(1, days_amount + 14):
        ws[get_column_letter(i) + "17"] = str(i)
        ws[get_column_letter(i) + "17"].alignment = Alignment(horizontal = "center")
    
def create_title():
    for x in range(5, 16):
        for y in range(5, 8):
            ws.cell(row=y, column=x).border = Border(bottom=thins)
    ws['G1'] = "Т а б е л ь  №  ______________"
    ws['G1'].font = bold_font

    ws['F2'] = "учета использования рабочего времени"
    ws['F2'].font = bold_font

    ws['B5'] = "Учреждение"
    ws['E5'] = "Ярославский государственный университет им. П.Г. Демидова"
    ws['B6'] = "Структурное подразделение"
    ws['E6'] = "Кафедра теоретической информатики"
    ws['B7'] = "Вид табеля"
    ws['E7'] = "первичный"

    ws['L8'] = "(первичный - 0; корректирующий - 1, 2 и так далее)"
    ws['H4'] = "за период с 1"
    ws['M4'] = "по"
    ws['N4'] = str(days_amount) + " " + months[date[0]]

    ws['C10'] = "учётный номер"
    ws['C10'].font = small_font
    ws['c10'].alignment = Alignment(vertical = "center")
    ws.merge_cells('F10:AH10')

    ws['F10'] = "Числа месяца"
    ws['F10'].alignment = Alignment(horizontal = "center")

    ws[get_column_letter(days_amount + 8) + "10"] = "Итого отработано за месяц"
    ws[get_column_letter(days_amount + 8) + "10"].alignment = Alignment(horizontal = "center")
    ws.merge_cells(get_column_letter(days_amount + 8) + "10" + ":" + get_column_letter(days_amount + 12) + "10")

    ws[get_column_letter(days_amount + 13) + "10"] = "Количе-ство дней (часов) неявок"
    ws[get_column_letter(days_amount + 13) + "10"].alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
    ws.merge_cells(get_column_letter(days_amount + 13) + "10" + ":" + get_column_letter(days_amount + 13) + "16")

    ws['B11'] = "Фамилия, имя"
    ws['B12'] = "Отчество"
    ws['U11'] = "Итого"

    ws[get_column_letter(days_amount + 7) + "11"] = "Всего"
    ws[get_column_letter(days_amount + 8) + "11"] = "часов"
    ws[get_column_letter(days_amount + 8) + "11"].alignment = Alignment(horizontal = "center", vertical = "center")
    ws.merge_cells(get_column_letter(days_amount + 8) + "11" + ":" + get_column_letter(days_amount + 12) + "12")


basic_formating()
create_title()
wb.save("sample.xlsx")