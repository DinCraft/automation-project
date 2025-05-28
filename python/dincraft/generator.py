import sys
from calendar import monthrange
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment

from dincraft.datetime.utils import Utils
from dincraft.datetime.month import Month
from dincraft.generator.cell import Cell
from dincraft.common.employee import Employee
from dincraft.common.table import Table

table = Table()
month = Month(3)
print(month.id(), month.name())
exit(0)

if len(sys.argv) != 2:
    exit("One argument is required!\n<date> in this format: 01/2020")

bold_font = Font(name='Arimo', size=10, bold=True)
regular_font = Font(name='Arimo', size=10)
small_font = Font(name='Arimo', size=8)
thins = Side(border_style="thin")

wb = Workbook()
Cell().init(wb.active)
ws = wb.active
date = sys.argv[1].split("/")
days_amount = monthrange(int(date[1]), int(date[0]))[1]
today = datetime.date.today()

def basic_formating():
    ws.row_dimensions[1].height = 70
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 18

    for i in range(1, 16):
        ws.column_dimensions[get_column_letter(5 + i)].width = 4
        ws[get_column_letter(5 + i) + "13"] = str(i)
        ws[get_column_letter(
            5 + i) + "15"] = Utils.get_day_of_week(int(date[1]), int(date[0]), i)
    for i in range(1, days_amount - 14):
        ws.column_dimensions[get_column_letter(21 + i)].width = 4
        ws[get_column_letter(21 + i) + "13"] = str(15 + i)
        ws[get_column_letter(
            21 + i) + "15"] = Utils.get_day_of_week(int(date[1]), int(date[0]), 15 + i)
    for i in range(1, days_amount + 2):
        ws[get_column_letter(5 + i) +
           "13"].alignment = Alignment(horizontal="center")
    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(7 + days_amount + i)].width = 10
    for i in range(1, days_amount + 14):
        ws[get_column_letter(i) + "17"] = str(i)
        ws[get_column_letter(i) + "17"].alignment = Alignment(horizontal="center")


def create_title():
    for x in range(5, 30):
        for y in range(5, 8):
            ws.cell(row=y, column=x).border = Border(bottom=thins)
    ws['G1'] = "Т а б е л ь  №  ______________"
    ws['G1'].font = bold_font

    ws['F2'] = "учета использования рабочего времени"
    ws['F2'].font = bold_font

    ws['B5'] = "Учреждение"
    ws['E5'] = "Ярославский государственный университет им. П.Г. Демидова"
    ws['B6'] = "Структурное подразделение"
    ws['E6'] = "Кафедра теоретической информатики"
    ws['B7'] = "Вид табеля"
    ws['E7'] = "первичный"

    ws['L8'] = "(первичный - 0; корректирующий - 1, 2 и так далее)"
    ws['H4'] = "за период с 1"
    ws['M4'] = "по"
    ws['N4'] = str(days_amount) + " " + Months.months[date[0]] + " " + date[1]

    ws.merge_cells('C10:D10')
    ws['C10'] = "учётный номер"
    ws['C10'].font = small_font
    ws['c10'].alignment = Alignment(vertical="center", horizontal="center")

    ws.merge_cells('F10:AH10')
    ws['F10'] = "Числа месяца"
    ws['F10'].alignment = Alignment(horizontal="center")

    ws[get_column_letter(days_amount + 8) + "10"] = "Итого отработано за месяц"
    ws[get_column_letter(days_amount + 8) + "10"].alignment = Alignment(
        horizontal="center")
    ws.merge_cells(get_column_letter(days_amount + 8) + "10" +
                   ":" + get_column_letter(days_amount + 12) + "10")

    ws[get_column_letter(days_amount + 13) +"10"] = "Количе-ство дней (часов) неявок"
    ws[get_column_letter(days_amount + 13) + "10"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True)
    ws.merge_cells(get_column_letter(days_amount + 13) + "10"
        + ":" + get_column_letter(days_amount + 13) + "16")

    ws['E13'] = "Должность"
    ws['E13'].alignment = Alignment(horizontal="center")
    ws['E14'] = "(профессия)"
    ws['E14'].alignment = Alignment(horizontal="center")
    ws['B11'] = "Фамилия, имя"
    ws['B11'].alignment = Alignment(horizontal="center")
    ws['B12'] = "отчество"
    ws['B12'].alignment = Alignment(horizontal="center")
    ws['A12'] = "№ п/п"
    ws['A12'].alignment = Alignment(horizontal="center")
    ws['U11'] = "Итого"
    ws['U11'].alignment = Alignment(horizontal="center")
    ws['U12'] = "дней"
    ws['U12'].alignment = Alignment(horizontal="center")
    ws['U13'] = "(часов)"
    ws['U13'].alignment = Alignment(horizontal="center")
    ws['U14'] = "явок"
    ws['U14'].alignment = Alignment(horizontal="center")
    ws['U15'] = "(неявок)"
    ws['U15'].alignment = Alignment(horizontal="center")
    ws['U16'] = "с 1 по 15"
    ws['U16'].alignment = Alignment(horizontal="center")

    ws[get_column_letter(days_amount + 7) + "11"] = "Всего"
    ws[get_column_letter(days_amount + 7) +
       "11"].alignment = Alignment(horizontal="center")
    ws[get_column_letter(days_amount + 7) + "12"] = "дней"
    ws[get_column_letter(days_amount + 7) +
       "12"].alignment = Alignment(horizontal="center")
    ws[get_column_letter(days_amount + 7) + "13"] = "(часов)"
    ws[get_column_letter(days_amount + 7) +
       "13"].alignment = Alignment(horizontal="center")
    ws[get_column_letter(days_amount + 7) + "14"] = "явок"
    ws[get_column_letter(days_amount + 7) +
       "14"].alignment = Alignment(horizontal="center")
    ws[get_column_letter(days_amount + 7) + "15"] = "(неявок)"
    ws[get_column_letter(days_amount + 7) +
       "15"].alignment = Alignment(horizontal="center")
    ws[get_column_letter(days_amount + 7) + "16"] = "за месяц"
    ws[get_column_letter(days_amount + 7) +
       "16"].alignment = Alignment(horizontal="center")

    ws[get_column_letter(days_amount + 8) + "11"] = "часов"
    ws[get_column_letter(days_amount + 8) + "11"].alignment = Alignment(
        horizontal="center", vertical="center")
    ws.merge_cells(get_column_letter(days_amount + 8) + "11" +
                   ":" + get_column_letter(days_amount + 12) + "12")

    ws[get_column_letter(days_amount + 8) + "13"] = "всего"
    ws[get_column_letter(days_amount + 8) + "13"].alignment = Alignment(
        horizontal="center",
        vertical="center")
    ws.merge_cells(get_column_letter(days_amount + 8) + "13" +
                   ":" + get_column_letter(days_amount + 8) + "16")

    ws[get_column_letter(days_amount + 9) + "13"] = "из них"
    ws[get_column_letter(days_amount + 9) + "13"].alignment = Alignment(
        horizontal="center",
        vertical="center")
    ws.merge_cells(get_column_letter(days_amount + 9) + "13" +
                   ":" + get_column_letter(days_amount + 12) + "13")

    ws[get_column_letter(days_amount + 9) + "14"] = "сверхурочных"
    ws[get_column_letter(days_amount + 9) + "14"].alignment = Alignment(
        horizontal="center", vertical="center", wrapText=True)
    ws.merge_cells(get_column_letter(days_amount + 9) + "14" +
                   ":" + get_column_letter(days_amount + 9) + "16")

    ws[get_column_letter(days_amount + 10) + "14"] = "ночных"
    ws[get_column_letter(days_amount + 10) + "14"].alignment = Alignment(
        horizontal="center", vertical="center", wrapText=True)
    ws.merge_cells(get_column_letter(days_amount + 10) + "14" +
                   ":" + get_column_letter(days_amount + 10) + "16")

    ws[get_column_letter(days_amount + 11) + "14"] = "выходных, праздничных"
    ws[get_column_letter(days_amount + 11) + "14"].alignment = Alignment(
        horizontal="center", vertical="center", wrapText=True)
    ws.merge_cells(get_column_letter(days_amount + 11) + "14" +
                   ":" + get_column_letter(days_amount + 11) + "16")


def add_metadata():
    # Коды
    ws.merge_cells('AK3:AM3')
    ws['AK3'] = "Коды"
    ws['AK3'].alignment = Alignment(horizontal="center")

    # Форма по ОКУД
    ws['AE4'] = "Форма по ОКУД"
    ws.merge_cells('AK4:AM4')
    ws['AK4'] = "0504421"
    ws['AK4'].alignment = Alignment(horizontal="center")

    # Дата
    ws['AE5'] = "Дата"

    _month = str(today.month)
    if len(_month) == 1:
        _month = "0" + _month
    ws['AK5'] = str(today.day) + "." + _month + "." + str(today.year)

    # По ОКПО
    ws['AE6'] = "По ОКПО"
    ws.merge_cells('AK6:AM6')
    ws['AK6'] = "02069409"
    ws['AK6'].alignment = Alignment(horizontal="center")

    # Номер корректировки
    ws['AE8'] = "Номер корректировки"
    ws['AK8'] = "0"

    # Дата формирования документа
    ws['AE9'] = "Дата формирования документа"
    ws['AK9'] = str(today.day) + "." + _month + "." + str(today.year)


basic_formating()
create_title()
add_metadata()

wb.save("sample.xlsx")
